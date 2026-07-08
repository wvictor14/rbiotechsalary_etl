"""Extract survey responses from the public Google Sheet into per-year raw CSVs.

Downloads the whole workbook as xlsx (one request), then writes one CSV per
year tab to data/raw/src_responses_<year>.csv with sanitized column headers.

Usage:
    uv run python scripts/extract.py
"""

import csv
import sys
import urllib.request
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from sanitize_headers import clean_name

SPREADSHEET_ID = "1G0FmJhkOME_sv66hWmhnZS5qR2KMTY7nzkxksv46bfk"
EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
)
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "raw"


def download_workbook(url=EXPORT_URL):
    """Download the workbook xlsx and return it loaded via openpyxl."""
    print(f"Downloading workbook from {url}")
    with urllib.request.urlopen(url) as response:
        if response.status != 200:
            raise RuntimeError(f"Download failed with HTTP {response.status}")
        payload = response.read()
    print(f"Downloaded {len(payload) / 1024:.1f} KiB")
    return openpyxl.load_workbook(BytesIO(payload), read_only=True)


def dedupe_names(names):
    """Append _2, _3, ... to repeated sanitized column names."""
    seen = {}
    result = []
    for name in names:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 1
            result.append(name)
    return result


def format_cell(value):
    """Render a cell as a CSV-safe string; DuckDB re-infers types on read."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return str(value)


def extract_sheet(worksheet, output_path):
    """Write one worksheet to CSV with sanitized, deduped headers."""
    rows = worksheet.iter_rows(values_only=True)
    raw_header = next(rows)

    # Drop columns with empty headers (trailing junk columns in some years)
    keep = [i for i, name in enumerate(raw_header) if name not in (None, "")]
    header = dedupe_names([clean_name(raw_header[i]) for i in keep])

    n_rows = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            # First column is the submission timestamp; skip empty rows
            if row[keep[0]] is None:
                continue
            writer.writerow([format_cell(row[i]) if i < len(row) else "" for i in keep])
            n_rows += 1

    print(f"  {output_path.name}: {n_rows} rows, {len(header)} columns")
    return n_rows


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = download_workbook()

    total = 0
    for worksheet in workbook.worksheets:
        year = worksheet.title.strip()
        if not year.isdigit():
            print(f"  skipping non-year sheet: {year!r}")
            continue
        total += extract_sheet(worksheet, OUTPUT_DIR / f"src_responses_{year}.csv")

    print(f"✅ Extracted {total} total rows to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
